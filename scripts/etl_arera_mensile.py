"""
ETL ARERA - Prezzi finali tutela elettricità e CMEMm gas.

Sorgente: ARERA (Autorità di Regolazione per Energia Reti e Ambiente)
Pubblicazione: 
  - Elettricità tutela vulnerabili: aggiornamento trimestrale 
    https://www.arera.it/dati-e-statistiche/dettaglio/aggiornamenti-delle-condizioni-di-tutela-elettricita
  - CMEMm gas (materia prima): aggiornamento mensile
    https://www.arera.it/area-operatori/prezzi-e-tariffe/valore-cmemm

Licenza: I dati ARERA sono pubblici. Citazione obbligatoria della fonte.

Output:
- Tab `prezzi_finali_arera` del Google Sheet master
- Tab `metadati_aggiornamento` con log ETL
"""

import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import requests
from loguru import logger

sys.path.insert(0, str(Path(__file__).parent))

from utils.gsheets_client import (
    get_gspread_client,
    open_master_sheet,
    get_or_create_worksheet,
    log_etl_run,
)


# COSTANTI

# URL diretti file XLS ARERA (verificati maggio 2026)
URL_TUTELA_ELETTRICITA = "https://www.arera.it/fileadmin/allegati/dati/ele/eep35new.xlsx"

# Pagina di partenza per ricerca CMEMm (l'URL del file XLS può variare ogni mese)
URL_PAGINA_CMEMM = "https://www.arera.it/area-operatori/prezzi-e-tariffe/valore-cmemm-servizio-di-tutela-della-vulnerabilita"

# Schema dati output
HEADERS_ARERA = [
    "anno_mese",          # YYYY-MM
    "tipo_dato",          # "elettricita_tutela" | "cmemm_gas"
    "periodo",            # "Q1 2026" | "Aprile 2026" | etc
    "valore",             # numero
    "unita",              # "cent/kWh" | "€/MWh"
    "fonte_url",
    "data_inserimento",   # timestamp ETL
]


def download_xlsx(url: str, label: str) -> bytes:
    """Scarica file XLSX da URL come bytes."""
    logger.info(f"Download {label}: {url}")
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    logger.info(f"  -> {len(response.content)} bytes scaricati")
    return response.content


def parse_tutela_elettricita(xlsx_bytes: bytes) -> list[dict]:
    """
    Parsa il file XLSX della tutela elettricità ARERA (eep35new.xlsx).

    Il file contiene la serie storica dei prezzi per cliente tipo (2700 kWh/anno).
    Cerchiamo le colonne con periodo (es. "1° trimestre 2026") e valore €/kWh.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    record_estratti = []

    # Itera tutti i fogli (di solito ce ne sono 2: 2700kWh e 2000kWh)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"  Foglio '{sheet_name}': {ws.max_row} righe x {ws.max_column} colonne")

        # DIAGNOSTICA: stampa le prime 5 righe per capire la struttura
        logger.info(f"  Prime righe del foglio:")
        for r in range(1, min(6, ws.max_row + 1)):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, min(8, ws.max_column + 1))]
            logger.info(f"    Riga {r}: {row_values}")

        # Logica estrazione: ARERA usa schemi che cambiano nel tempo.
        # Strategia: cerchiamo tutte le celle con valori numerici tra 10 e 100
        # (centesimi/kWh ragionevoli) accompagnate da etichette di periodo.
        for row in ws.iter_rows(values_only=True):
            for i, cell_value in enumerate(row):
                if cell_value is None:
                    continue
                # Cerca una stringa che indichi un trimestre o periodo
                if isinstance(cell_value, str):
                    cell_lower = cell_value.lower()
                    if "trimestre" in cell_lower or "trim" in cell_lower:
                        # Le celle vicine potrebbero contenere il valore numerico
                        for j in range(max(0, i - 2), min(len(row), i + 5)):
                            v = row[j]
                            if isinstance(v, (int, float)) and 5 < v < 100:
                                record_estratti.append({
                                    "periodo": cell_value.strip(),
                                    "valore": round(float(v), 4),
                                    "foglio": sheet_name,
                                })
                                break

    logger.info(f"  Record estratti tutela elettricità: {len(record_estratti)}")
    return record_estratti


def parse_cmemm_gas(xlsx_bytes: bytes) -> list[dict]:
    """
    Parsa il file XLSX del CMEMm gas ARERA.
    Schema atteso: tabella con mese | valore_eur_mwh
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    record_estratti = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"  Foglio '{sheet_name}': {ws.max_row} righe x {ws.max_column} colonne")

        logger.info(f"  Prime righe del foglio:")
        for r in range(1, min(6, ws.max_row + 1)):
            row_values = [ws.cell(row=r, column=c).value for c in range(1, min(6, ws.max_column + 1))]
            logger.info(f"    Riga {r}: {row_values}")

        # Strategia: cerchiamo celle con date (datetime) o nomi mesi (es. "gennaio 2026")
        # accompagnate da numeri 10-150 €/MWh ragionevoli per il CMEMm
        mesi_italiani = ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                         "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

        for row in ws.iter_rows(values_only=True):
            for i, cell_value in enumerate(row):
                if cell_value is None:
                    continue

                periodo = None

                if isinstance(cell_value, datetime):
                    periodo = cell_value.strftime("%Y-%m")
                elif isinstance(cell_value, str):
                    cell_lower = cell_value.lower()
                    for mese in mesi_italiani:
                        if mese in cell_lower:
                            periodo = cell_value.strip()
                            break

                if periodo:
                    for j in range(max(0, i - 2), min(len(row), i + 5)):
                        v = row[j]
                        if isinstance(v, (int, float)) and 5 < v < 200:
                            record_estratti.append({
                                "periodo": periodo,
                                "valore": round(float(v), 4),
                                "foglio": sheet_name,
                            })
                            break

    logger.info(f"  Record estratti CMEMm gas: {len(record_estratti)}")
    return record_estratti


def estrai_anno_mese(periodo_str: str) -> str:
    """
    Converte stringhe di periodo varie in formato YYYY-MM.
    Esempi:
      "1° trimestre 2026" -> "2026-01" (primo mese del trimestre)
      "Aprile 2026" -> "2026-04"
      "2026-04" -> "2026-04"
    """
    s = periodo_str.lower().strip()

    # Già nel formato YYYY-MM
    if len(s) == 7 and s[4] == "-":
        return s

    mesi_map = {
        "gennaio": "01", "febbraio": "02", "marzo": "03", "aprile": "04",
        "maggio": "05", "giugno": "06", "luglio": "07", "agosto": "08",
        "settembre": "09", "ottobre": "10", "novembre": "11", "dicembre": "12",
    }
    trim_map = {"1": "01", "2": "04", "3": "07", "4": "10"}

    anno = None
    mese = None
    for word in s.split():
        if word.isdigit() and len(word) == 4:
            anno = word
        if word in mesi_map:
            mese = mesi_map[word]
        if "°" in word:
            num = word.replace("°", "").strip()
            if num in trim_map:
                mese = trim_map[num]

    if anno and mese:
        return f"{anno}-{mese}"
    return ""


def main():
    logger.info("=" * 60)
    logger.info("ETL ARERA - Avvio")
    logger.info("=" * 60)

    spreadsheet = None
    record_caricati = 0
    esito = "ok"
    note = ""

    try:
        client = get_gspread_client()
        spreadsheet = open_master_sheet(client)

        worksheet = get_or_create_worksheet(
            spreadsheet,
            "prezzi_finali_arera",
            headers=HEADERS_ARERA,
            rows=2000,
        )

        # 1. ELETTRICITÀ TUTELA
        try:
            xlsx_ele = download_xlsx(URL_TUTELA_ELETTRICITA, "tutela elettricità")
            record_ele = parse_tutela_elettricita(xlsx_ele)
        except Exception as e:
            logger.error(f"Errore elettricità: {e}")
            record_ele = []

        # 2. CMEMm GAS — Per ora skippato perché URL del file cambia ogni mese
        # Sarà implementato in versione successiva con scraping della pagina
        logger.warning("CMEMm gas: implementazione URL dinamico in fase successiva")
        record_gas = []

        # Compongo le righe da scrivere
        timestamp = datetime.utcnow().isoformat(timespec="seconds")
        righe = []
        for r in record_ele:
            anno_mese = estrai_anno_mese(r["periodo"])
            if not anno_mese:
                continue
            righe.append([
                anno_mese,
                "elettricita_tutela",
                r["periodo"],
                r["valore"],
                "cent/kWh",
                URL_TUTELA_ELETTRICITA,
                timestamp,
            ])

        for r in record_gas:
            anno_mese = estrai_anno_mese(r["periodo"])
            if not anno_mese:
                continue
            righe.append([
                anno_mese,
                "cmemm_gas",
                r["periodo"],
                r["valore"],
                "€/MWh",
                URL_PAGINA_CMEMM,
                timestamp,
            ])

        # Deduplica per (anno_mese, tipo_dato) tenendo l'ultimo valore
        if righe:
            df = pd.DataFrame(righe, columns=HEADERS_ARERA)
            df = df.drop_duplicates(subset=["anno_mese", "tipo_dato"], keep="last")
            df = df.sort_values(["tipo_dato", "anno_mese"]).reset_index(drop=True)

            # Sovrascrivo l'intero tab (modalità bulk per evitare rate limit)
            contenuto = [HEADERS_ARERA] + df.fillna("").astype(str).values.tolist()
            worksheet.clear()
            worksheet.update(
                values=contenuto,
                range_name="A1",
                value_input_option="USER_ENTERED",
            )

            record_caricati = len(df)
            logger.info(f"Scritti {record_caricati} record nel tab prezzi_finali_arera")
            note = f"Elettricità: {len(record_ele)} record. Gas: {len(record_gas)} record."
        else:
            esito = "errore"
            note = "Nessun record estratto"
            logger.warning(note)

    except Exception as e:
        logger.exception("Errore durante ETL ARERA")
        esito = "errore"
        note = str(e)[:500]
        raise

    finally:
        if spreadsheet:
            try:
                log_etl_run(
                    spreadsheet=spreadsheet,
                    fonte="ARERA",
                    record_caricati=record_caricati,
                    esito=esito,
                    url_fonte=URL_TUTELA_ELETTRICITA,
                    note=note,
                )
            except Exception:
                logger.error("Impossibile loggare l'esito ETL")

    logger.info("=" * 60)
    logger.info(f"ETL ARERA - Fine. Esito: {esito}, record: {record_caricati}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
